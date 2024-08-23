
import os

if os.getenv('LAMBDA_ENV') == None:
    from openpyxl.utils import get_column_letter
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
    import sys
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference, Series
    from openpyxl.chart.label import DataLabelList
    from openpyxl.drawing.image import Image

def test_set_statistics(workbook,PredictionResults,test_set_names,y_actuals):
    """
    Creates the table of test set statistics on the spreadsheet
    """
    #Make sure y_actuals is a list
    if not isinstance(y_actuals,list):
        y_actuals = y_actuals.tolist()
        
    #Set up variables and column names
    name = test_set_names
    yhat = PredictionResults.yhat
    yhat_linear = PredictionResults.y_linear
    phi = PredictionResults.phi
    adjusted_fit = PredictionResults.adjusted_fit
    fit = PredictionResults.fit
    yhat_compound = PredictionResults.yhat_compound
    adjusted_fit_compound = PredictionResults.adjusted_fit_compound
    fit_compound = PredictionResults.fit_compound
    y_actuals = y_actuals
    
    #Make a dataframe of the data going to the sheet
    df = {
          'names':name,
          'yhat':yhat,
          'yhat_linear':yhat_linear,
          'phi':phi,
          'adjusted_fit':adjusted_fit,
          'fit':fit,
          'yhat_compound':yhat_compound,
          'adjusted_fit_compound':adjusted_fit_compound,
          'fit_compound':fit_compound,
          'y_actuals':y_actuals
            }
    df = pd.DataFrame(df)
    
    #Add the dataframe to the sheet
    sheet = workbook.create_sheet(title='Test Set Statistics')
    thin = Side(style='thin')
    allborder = Border(left=thin,right=thin,top=thin,bottom=thin)
    bold_font = Font(bold=True)
    for c_idx, column_name in enumerate(df.columns, 1):
        cell = sheet.cell(row=1, column=c_idx, value=column_name)
        cell.border = allborder  
        cell.font = bold_font    
        
    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            cell.border = allborder
            cell.number_format = '0.00'
            if c_idx == 1:
                cell.font = bold_font
            
    #center everything
    alignment = Alignment(horizontal='center', vertical='center')
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = alignment

    #no gridlines
    sheet.sheet_view.showGridLines = False
    
    #autofit column width
    for col in range(1, 12):  
        column_letter = get_column_letter(col)
        max_length = 0
        for row in range(1, sheet.max_row + 1):
                cell_value = sheet[f'{column_letter}{row}'].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
        adjusted_width = (max_length) * 1.2  # Adjusting factor for better fit
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    
