
import os

if os.getenv('LAMBDA_ENV') == None:
    from openpyxl.utils import get_column_letter
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.drawing.image import Image
    from .create_graphs import *
    from .create_table import *


# Function to create the y_actual_means sheet
def create_y_actual_means(workbook,df, result_path,high_percentile_value=0.80, low_percentile_value=0.20):
    """
    Create y actual means table
    """
    #round everything to 2 decimal points
    df = df.round(2)
    
    #Create a new sheet
    sheet = workbook.create_sheet(title='y_actual_means')
    
    #Set thick/thin border styles
    thin = Side(style='thin')
    thick = Side(style='thick')
    
    #Create big header
    title = 'y_actual Mean'
    merged_cell_high = sheet['B2']
    merged_cell_high.value = title
    merged_cell_high.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell_high.fill = PatternFill(start_color='00B0F0',end_color='00B0F0',fill_type='solid')
    merged_cell_high.font = Font(color='FFFFFF',bold=True)
    merged_cell_high.border = Border(top=thick, left=thick, right=thick, bottom=thin)
    sheet.merge_cells('B2:F2')
    
    #Write column headers with specified formatting
    fill_color = PatternFill(start_color = 'D9D9D9', end_color='D9D9D9',fill_type='solid')
    sheet['B3'] = 'Sample'
    sheet['C3'] = 'Sub Sample'
    sheet['D3'] = 'y_actual Mean'
    sheet['E3'] = 'yhat Cutoff'
    sheet['F3'] = 'fit Cutoff'
    headers = ['B3','C3','D3','E3','F3']
    for cell in headers:
        sheet[cell].fill = fill_color

    # Create full sample row
    sheet['B4'] = "Full"
    sheet['C4'] = "Full"
    full_sample = df.loc['Full Sample'].values[0]
    sheet['D4'] = full_sample 
    for col in list('BCD'):
        sheet[f'{col}4'].fill = PatternFill(start_color='FDE9D9', end_color='FDE9D9', fill_type='solid')

    #Create "When Yhat is high" block
    merged_cell_high = sheet['B5']
    merged_cell_high.value = "When yhat is high"
    merged_cell_high.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell_high.border = Border(left=thick, top=thin, bottom=thin, right=thin)
    sheet.merge_cells('B5:B7')
    sheet['C5'] = "Full"
    sheet['C6'] = "High Fit"
    sheet['C7'] = "Low Fit"
    high_prediction = df.loc['High Prediction'].values[0]
    sheet['D5'] = high_prediction
    high_prediction_w_high_fit = df.loc['High Prediction w/ High Fit'].values[0]
    sheet['D6'] = high_prediction_w_high_fit
    high_prediction_w_low_fit = df.loc['High Prediction w/ Low Fit'].values[0]
    sheet['D7'] = high_prediction_w_low_fit
    merged_cell_high = sheet['E5']
    merged_cell_high.value = high_percentile_value
    merged_cell_high.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell_high.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    sheet.merge_cells('E5:E7')
    sheet['F6'] = high_percentile_value
    sheet['F7'] = low_percentile_value
    for row in range(5, 8):
        for col in list('BCDEF'):
            sheet[f'{col}{row}'].fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
            
    #Create "When yhat is low" block
    merged_cell_low = sheet['B8']
    merged_cell_low.value = "When yhat is low"
    merged_cell_low.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell_low.border = Border(left=thick, bottom=thick, top=thin, right=thin)
    sheet.merge_cells('B8:B10')
    sheet['C8'] = "Full"
    sheet['C9'] = "High Fit"
    sheet['C10'] = "Low Fit"
    low_prediction = df.loc['Low Prediction'].values[0]
    sheet['D8'] = low_prediction
    low_prediction_w_high_fit = df.loc['Low Prediction w/ High Fit'].values[0]
    sheet['D9'] = low_prediction_w_high_fit
    low_prediction_w_low_fit = df.loc['Low Prediction w/ Low Fit'].values[0]
    sheet['D10'] = low_prediction_w_low_fit
    merged_cell_high = sheet['E8']
    merged_cell_high.value = low_percentile_value
    merged_cell_high.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell_high.border = Border(top=thin, bottom=thick, left=thin, right=thin)
    sheet.merge_cells('E8:E10')
    sheet['F9'] = high_percentile_value
    sheet['F10'] = low_percentile_value
    for row in range(8, 11):
        for col in list('BCDEF'):
            sheet[f'{col}{row}'].fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
            
    #set black backgrounds for cells
    black_background_cells = ['E4','F4','F5','F8']
    for cell in black_background_cells:
        sheet[cell].fill = PatternFill(start_color='000000',end_color='000000',fill_type='solid')

    #set borders
    # B3
    B3_border = Border(left=thick, top=thin, bottom=thin, right=thin)
    sheet['B3'].border = B3_border

    # B4
    B4_border = Border(left=thick, top=thin, bottom=thin, right=thin)
    sheet['B4'].border = B4_border
    
    # C3
    C3_border = Border(top=thin, bottom=thin, left=thin, right=thin)
    sheet['C3'].border = C3_border

    # C4
    C4_border = Border(top=thin, bottom=thin, left=thin, right=thin)
    sheet['C4'].border = C4_border

    # C5
    C5_border = Border(top=thin, left=thin, right=thin)
    sheet['C5'].border = C5_border

    # C6
    C6_border = Border(left=thin, right=thin)
    sheet['C6'].border = C6_border

    # C7
    C7_border = Border(bottom=thin, left=thin, right=thin)
    sheet['C7'].border = C7_border

    # C8
    C8_border = Border(top=thin, left=thin, right=thin)
    sheet['C8'].border = C8_border

    # C9
    C9_border = Border(left=thin, right=thin)
    sheet['C9'].border = C9_border

    # C10
    C10_border = Border(bottom=thick, left=thin, right=thin)
    sheet['C10'].border = C10_border

    # D3
    D3_border = Border(top=thin, bottom=thin, left=thin, right=thin)
    sheet['D3'].border = D3_border

    # D4
    D4_border = Border(top=thin, bottom=thin, left=thin, right=thin)
    sheet['D4'].border = D4_border

    # D5
    D5_border = Border(top=thin, left=thin, right=thin)
    sheet['D5'].border = D5_border

    # D6
    D6_border = Border(left=thin, right=thin)
    sheet['D6'].border = D6_border

    # D7
    D7_border = Border(bottom=thin, left=thin, right=thin)
    sheet['D7'].border = D7_border

    # D8
    D8_border = Border(top=thin, left=thin, right=thin)
    sheet['D8'].border = D8_border

    # D9
    D9_border = Border(left=thin, right=thin)
    sheet['D9'].border = D9_border

    # D10
    D10_border = Border(bottom=thick, left=thin, right=thin)
    sheet['D10'].border = D10_border

    # E3
    E3_border = Border(top=thin, bottom=thin, left=thin, right=thin)
    sheet['E3'].border = E3_border

    # F3
    F3_border = Border(top=thin, bottom=thin, left=thin, right=thick)
    sheet['F3'].border = F3_border

    # F4
    F4_border = Border(top=thin, bottom=thin, left=thin, right=thick)
    sheet['F4'].border = F4_border

    # F5
    F5_border = Border(top=thin,left=thin, right=thick)
    sheet['F5'].border = F5_border

    # F6
    F6_border = Border(left=thin, right=thick)
    sheet['F6'].border = F6_border

    # F7
    F7_border = Border(bottom=thin, left=thin, right=thick)
    sheet['F7'].border = F7_border

    # F8
    F8_border = Border(top=thin,left=thin, right=thick)
    sheet['F8'].border = F8_border

    # F9
    F9_border = Border(left=thin, right=thick)
    sheet['F9'].border = F9_border

    # F10
    F10_border = Border(bottom=thick, left=thin, right=thick)
    sheet['F10'].border = F10_border
    
    #center everything
    alignment = Alignment(horizontal='center', vertical='center')
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = alignment

    #no gridlines
    sheet.sheet_view.showGridLines = False
    
    #autofit column width
    for col in range(2, 7):  # Columns B to F
        column_letter = get_column_letter(col)
        max_length = 0
        for row in range(1, sheet.max_row + 1):
                cell_value = sheet[f'{column_letter}{row}'].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
        adjusted_width = (max_length) * 1.2  # Adjusting factor for better fit
        sheet.column_dimensions[column_letter].width = adjusted_width

    #autofit row height
    for row in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 15  # Adjust the height as needed

    #format to 2 decimal places
    numeric_cells = ['D4','D5','D6','D7','D8','D9','D10','D11','D12']
    
    for cell in numeric_cells:
        sheet[cell].number_format = '0.00'
    
    #add in plot image
    graph_path = y_actual_means_graph(df,result_path)
    img = Image(graph_path)
    sheet.add_image(img,'H1')

# Create the co-occurrence sheet
def create_co_occurrence(workbook, df, result_path, high_percentile_value=0.80, low_percentile_value=0.20):
    """
    Creates info weighted co-occurrence table
    """
    #round everything to 2 decimal points
    df = df.round(2)
    
    #create the sheet
    sheet = workbook.create_sheet(title='info_weighted_co_occurrence')

    #Set thick/thin border styles
    thin = Side(style='thin')
    thick = Side(style='thick')

    #Create big header
    title = 'Informativeness Weighted Co-Occurrence'
    merged_cell_high = sheet['B2']
    merged_cell_high.value = title
    merged_cell_high.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell_high.fill = PatternFill(start_color='7030A0',end_color='7030A0',fill_type='solid')
    merged_cell_high.font = Font(color='FFFFFF',bold=True)
    merged_cell_high.border = Border(top=thick, left=thick, right=thick, bottom=thin)
    sheet.merge_cells('B2:F2')
    
    #Write column headers with specified formatting
    fill_color = PatternFill(start_color = 'D9D9D9', end_color='D9D9D9',fill_type='solid')
    sheet['B3'] = 'Sample'
    sheet['C3'] = 'Sub Sample'
    sheet['D3'] = 'c(yhat, y_actual)'
    sheet['E3'] = 'yhat Cutoff'
    sheet['F3'] = 'fit Cutoff'
    headers = ['B3','C3','D3','E3','F3']
    for cell in headers:
        sheet[cell].fill = fill_color

    # Create "Full sample" block
    merged_cell_high = sheet['B4']
    merged_cell_high.value = "Full"
    merged_cell_high.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell_high.border = Border(left=thick, top=thin, bottom=thin, right=thin)
    sheet.merge_cells('B4:B6')
    sheet['C4'] = "Full"
    sheet['C5'] = "High Fit"
    sheet['C6'] = "Low Fit"
    full_sample = df.loc['Full Sample'].values[0]
    sheet['D4'] = full_sample
    high_fit = df.loc['High Fit'].values[0]
    sheet['D5'] = high_fit
    low_fit = df.loc['Low Fit'].values[0]
    sheet['D6'] = low_fit
    sheet['F5'] = high_percentile_value
    sheet['F6'] = low_percentile_value
    for row in range(4, 7):
        for col in list('BCDEF'):
            sheet[f'{col}{row}'].fill = PatternFill(start_color='FDE9D9', end_color='FDE9D9', fill_type='solid')

    #Create "When Yhat is high" block
    merged_cell_high = sheet['B7']
    merged_cell_high.value = "When yhat is high"
    merged_cell_high.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell_high.border = Border(left=thick, top=thin, bottom=thin, right=thin)
    sheet.merge_cells('B7:B9')
    sheet['C7'] = "Full"
    sheet['C8'] = "High Fit"
    sheet['C9'] = "Low Fit"
    high_prediction = df.loc['High Prediction'].values[0]
    sheet['D7'] = high_prediction
    high_prediction_w_high_fit = df.loc['High Prediction w/ High Fit'].values[0]
    sheet['D8'] = high_prediction_w_high_fit
    high_prediction_w_low_fit = df.loc['High Prediction w/ Low Fit'].values[0]
    sheet['D9'] = high_prediction_w_low_fit
    merged_cell_high = sheet['E7']
    merged_cell_high.value = high_percentile_value
    merged_cell_high.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell_high.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    sheet.merge_cells('E7:E9')
    sheet['F8'] = high_percentile_value
    sheet['F9'] = low_percentile_value
    for row in range(7, 10):
        for col in list('BCDEF'):
            sheet[f'{col}{row}'].fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
            
    #Create "When yhat is low" block
    merged_cell_low = sheet['B10']
    merged_cell_low.value = "When yhat is low"
    merged_cell_low.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell_low.border = Border(left=thick, bottom=thick, top=thin, right=thin)
    sheet.merge_cells('B10:B12')
    sheet['C10'] = "Full"
    sheet['C11'] = "High Fit"
    sheet['C12'] = "Low Fit"
    low_prediction = df.loc['Low Prediction'].values[0]
    sheet['D10'] = low_prediction
    low_prediction_w_high_fit = df.loc['Low Prediction w/ High Fit'].values[0]
    sheet['D11'] = low_prediction_w_high_fit
    low_prediction_w_low_fit = df.loc['Low Prediction w/ Low Fit'].values[0]
    sheet['D12'] = low_prediction_w_low_fit
    merged_cell_high = sheet['E10']
    merged_cell_high.value = low_percentile_value
    merged_cell_high.alignment = Alignment(horizontal='center', vertical='center')
    merged_cell_high.border = Border(top=thin, bottom=thick, left=thin, right=thin)
    sheet.merge_cells('E10:E12')
    sheet['F11'] = high_percentile_value
    sheet['F12'] = low_percentile_value
    for row in range(10, 13):
        for col in list('BCDEF'):
            sheet[f'{col}{row}'].fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
            
    #set black backgrounds for cells
    black_background_cells = ['E4','E5','E6','F4','F7','F10']
    for cell in black_background_cells:
        sheet[cell].fill = PatternFill(start_color='000000',end_color='000000',fill_type='solid')

    #set borders
    # B3
    B3_border = Border(left=thick, top=thin, bottom=thin, right=thin)
    sheet['B3'].border = B3_border
    
    # C3
    C3_border = Border(top=thin, bottom=thin, left=thin, right=thin)
    sheet['C3'].border = C3_border
    
    # D3
    D3_border = Border(top=thin, bottom=thin, left=thin, right=thin)
    sheet['D3'].border = D3_border
    
    # E3
    E3_border = Border(top=thin, bottom=thin, left=thin, right=thin)
    sheet['E3'].border = E3_border
    
    # F3
    F3_border = Border(top=thin, bottom=thin, left=thin, right=thick)
    sheet['F3'].border = F3_border
    
    # C4
    C4_border = Border(left=thin,top=thin,right=thin)
    sheet['C4'].border = C4_border
    
    # C5
    C5_border = Border(left=thin,right=thin)
    sheet['C5'].border = C5_border
    
    # C6
    C6_border = Border(left=thin,bottom=thin,right=thin)
    sheet['C6'].border = C6_border

    # C7
    C7_border = Border(top=thin, left=thin, right=thin)
    sheet['C7'].border = C7_border

    # C8
    C8_border = Border(left=thin, right=thin)
    sheet['C8'].border = C8_border

    # C9
    C9_border = Border(bottom=thin, left=thin, right=thin)
    sheet['C9'].border = C9_border

    # C10
    C10_border = Border(top=thin, left=thin, right=thin)
    sheet['C10'].border = C10_border

    # C11
    C11_border = Border(left=thin, right=thin)
    sheet['C11'].border = C11_border

    # C12
    C12_border = Border(bottom=thick, left=thin, right=thin)
    sheet['C12'].border = C12_border

    # D4
    D4_border = Border(left=thin,top=thin,right=thin)
    sheet['D4'].border = D4_border
    
    # D5
    D5_border = Border(left=thin,right=thin)
    sheet['D5'].border = D5_border
    
    # D6
    D6_border = Border(left=thin,bottom=thin,right=thin)
    sheet['D6'].border = D6_border
    
    # D7
    D7_border = Border(top=thin, left=thin, right=thin)
    sheet['D7'].border = D7_border

    # D8
    D8_border = Border(left=thin, right=thin)
    sheet['D8'].border = D8_border

    # D9
    D9_border = Border(bottom=thin, left=thin, right=thin)
    sheet['D9'].border = D9_border

    # D10
    D10_border = Border(top=thin, left=thin, right=thin)
    sheet['D10'].border = D10_border

    # D11
    D11_border = Border(left=thin, right=thin)
    sheet['D11'].border = D11_border

    # D12
    D12_border = Border(bottom=thick, left=thin, right=thin)
    sheet['D12'].border = D12_border

    # F5
    F5_border = Border(top=thin, left=thin, right=thick)
    sheet['F5'].border = F5_border

    # F6
    F6_border = Border(bottom=thin, left=thin, right=thick)
    sheet['F6'].border = F6_border

    # F8
    F8_border = Border(top=thin,left=thin, right=thick)
    sheet['F8'].border = F8_border

    # F9
    F9_border = Border(left=thin, right=thick,bottom=thick)
    sheet['F9'].border = F9_border

    # F11
    F11_border = Border(top=thin, left=thin, right=thick)
    sheet['F11'].border = F11_border
    
    # F12
    F12_border = Border(left=thin,right=thick,bottom=thick)
    sheet['F12'].border = F12_border
    
    # F4
    F4_border = Border(right=thick)
    sheet['F4'].border = F4_border
    
    # F7
    F7_border = Border(right=thick)
    sheet['F7'].border = F7_border

    # F10
    F10_border = Border(right=thick)
    sheet['F10'].border = F10_border    
    
    #center everything
    alignment = Alignment(horizontal='center', vertical='center')
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = alignment

    #no gridlines
    sheet.sheet_view.showGridLines = False
    
    #autofit column width
    for col in range(2, 7):  # Columns B to F
        column_letter = get_column_letter(col)
        max_length = 0
        for row in range(1, sheet.max_row + 1):
                cell_value = sheet[f'{column_letter}{row}'].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
        adjusted_width = (max_length) * 1.2  # Adjusting factor for better fit
        sheet.column_dimensions[column_letter].width = adjusted_width

    #autofit row height
    for row in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 15  # Adjust the height as needed

    #format to 2 decimal places
    numeric_cells = ['D4','D5','D6','D7','D8','D9','D10','D11','D12']
    
    for cell in numeric_cells:
        sheet[cell].number_format = '0.00'

    #add plot image in
    graph_path = info_weighted_co_occurrence_graph(df,result_path)
    img = Image(graph_path)
    sheet.add_image(img,'H3')

# Function to create the betas and t statistics sheet
def betas_and_t_statistics(workbook, df):
    """
    Creates table of betas and t statistics
    """
    #round everything to 2 decimal points
    df = df.round(2)
    
    #Create sheet
    sheet = workbook.create_sheet(title='betas_and_t_statistics')
    
    #Set thick/thin border styles
    thin = Side(style='thin')
    thick = Side(style='thick')
    
    #Create headers and subheaders  
    merge_list = ['C2:D2','E2:F2','G2:H2','I2:J2']
    header_list = ['Full Sample','High Fit','Mid Fit','Low Fit']
    color_list = ['FDE9D9','C5D9F1','E4DFEC','EBF1DE']
    for i in range(0,4):
        merged_cell_high = sheet[merge_list[i][:2]]
        merged_cell_high.value = header_list[i]
        merged_cell_high.alignment = Alignment(horizontal='center',vertical='center')
        merged_cell_high.font = Font(color='000000',bold=True)
        merged_cell_high.border = Border(top=thick,left=thin,right=thin)
        merged_cell_high.fill = PatternFill(start_color = color_list[i], end_color=color_list[i],fill_type='solid')
        if merge_list[i][:2] == 'I2':
            merged_cell_high.border = Border(top=thick,left=thin,right=thick)
        sheet.merge_cells(merge_list[i])
        number = int(merge_list[i][:2][1]) + 1
        left_under = f"{merge_list[i][0]}{number}"
        right_under = f"{merge_list[i][3]}{number}"
        sheet[left_under].value = 'Linear Regression'
        sheet[right_under].value = 'Excess RBP'
        sheet[left_under].fill = PatternFill(start_color=color_list[i],end_color=color_list[i],fill_type='solid')
        sheet[right_under].fill = PatternFill(start_color=color_list[i],end_color=color_list[i],fill_type='solid')
        sheet[left_under].border = Border(left=thin,bottom=thin)
        sheet[right_under].border = Border(right=thin,bottom=thin)
        if right_under == 'J3':
            sheet[right_under].border = Border(right=thick,bottom=thin)
    
    sheet['B2'].fill = PatternFill(start_color='BFBFBF',end_color='BFBFBF',fill_type='solid') 
    sheet['B2'].border = Border(left=thick,top=thick,right=thin)
    sheet['B3'].fill = PatternFill(start_color='BFBFBF',end_color='BFBFBF',fill_type='solid')
    sheet['B3'].border = Border(left=thick,bottom=thin,right=thin)
    sheet['B4'].value = 'Beta Coefficient'
    sheet['B4'].font = Font(color='000000',bold=True)
    sheet['B4'].fill = PatternFill(start_color='BFBFBF',end_color='BFBFBF',fill_type='solid')
    sheet['B4'].border = Border(left=thick,top=thin,right=thin)
    sheet['B5'].value = 't-Statistic'
    sheet['B5'].font = Font(color='000000')
    sheet['B5'].fill = PatternFill(start_color='BFBFBF',end_color='BFBFBF',fill_type='solid')
    sheet['B5'].border = Border(left=thick,bottom=thick,right=thin)
        
    #Fill in statistics for each thing    
    start_row = 4
    start_col = 3
    for i in range(df.shape[0]):
        for j in range(df.shape[1]):
            cell = sheet.cell(row = start_row + i, column = start_col + j)
            cell.value = df.iat[i,j]
    
    #Cell formatting
    # C4
    sheet['C4'].fill = PatternFill(start_color='FDE9D9',end_color='FDE9D9',fill_type='solid')
    sheet['C4'].border = Border(top=thin,left=thin)
    
    # C5
    sheet['C5'].fill = PatternFill(start_color='FDE9D9',end_color='FDE9D9',fill_type='solid')
    sheet['C5'].border = Border(bottom=thick,left=thin)
    
    # D4
    sheet['D4'].fill = PatternFill(start_color='FDE9D9',end_color='FDE9D9',fill_type='solid')
    sheet['D4'].border = Border(top=thin,right=thin)
    
    # D5
    sheet['D5'].fill = PatternFill(start_color='FDE9D9',end_color='FDE9D9',fill_type='solid')
    sheet['D5'].border = Border(bottom=thick,right=thin)
    
    # E4
    sheet['E4'].fill = PatternFill(start_color='C5D9F1',end_color='C5D9F1',fill_type='solid')
    sheet['E4'].border = Border(top=thin,left=thin)
    
    # E5
    sheet['E5'].fill = PatternFill(start_color='C5D9F1',end_color='C5D9F1',fill_type='solid')
    sheet['E5'].border = Border(bottom=thick,left=thin)
    
    # F4
    sheet['F4'].fill = PatternFill(start_color='C5D9F1',end_color='C5D9F1',fill_type='solid')
    sheet['F4'].border = Border(top=thin,right=thin)
    
    # F5
    sheet['F5'].fill = PatternFill(start_color='C5D9F1',end_color='C5D9F1',fill_type='solid')
    sheet['F5'].border = Border(bottom=thick,right=thin)
    
    # G4
    sheet['G4'].fill = PatternFill(start_color='E4DFEC',end_color='E4DFEC',fill_type='solid')
    sheet['G4'].border = Border(top=thin,left=thin)
    
    # G5
    sheet['G5'].fill = PatternFill(start_color='E4DFEC',end_color='E4DFEC',fill_type='solid')
    sheet['G5'].border = Border(bottom=thick,left=thin)
    
    # H4
    sheet['H4'].fill = PatternFill(start_color='E4DFEC',end_color='E4DFEC',fill_type='solid')
    sheet['H4'].border = Border(top=thin,right=thin)
    
    # H5
    sheet['H5'].fill = PatternFill(start_color='E4DFEC',end_color='E4DFEC',fill_type='solid')
    sheet['H5'].border = Border(bottom=thick,right=thin)
    
    # I4
    sheet['I4'].fill = PatternFill(start_color='EBF1DE',end_color='EBF1DE',fill_type='solid')
    sheet['I4'].border = Border(top=thin,left=thin)
    
    # I5
    sheet['I5'].fill = PatternFill(start_color='EBF1DE',end_color='EBF1DE',fill_type='solid')
    sheet['I5'].border = Border(bottom=thick,left=thin)
    
    # J4
    sheet['J4'].fill = PatternFill(start_color='EBF1DE',end_color='EBF1DE',fill_type='solid')
    sheet['J4'].border = Border(top=thin,right=thick)
    
    # J5
    sheet['J5'].fill = PatternFill(start_color='EBF1DE',end_color='EBF1DE',fill_type='solid')
    sheet['J5'].border = Border(bottom=thick,right=thick)
    
    #center everything
    alignment = Alignment(horizontal='center', vertical='center')
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = alignment

    #no gridlines
    sheet.sheet_view.showGridLines = False
    
    #autofit column width
    for col in range(2, 11):  # Columns B to F
        column_letter = get_column_letter(col)
        max_length = 0
        for row in range(1, sheet.max_row + 1):
                cell_value = sheet[f'{column_letter}{row}'].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
        adjusted_width = (max_length) * 1.2  # Adjusting factor for better fit
        sheet.column_dimensions[column_letter].width = adjusted_width

    #autofit row height
    for row in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 15  # Adjust the height as needed

    #format to 2 decimal places
    numeric_cells = ['C4','C5','D4','D5','E4','E5','F4','F5','G4','G5','H4','H5','I4','I5','J4','J5']
    
    for cell in numeric_cells:
        sheet[cell].number_format = '0.00'

def variable_importance(workbook,df,result_path):
    """
    Creates the variable importance table.
    """
    #round everything to 2 decimal points
    df = df.round(3)
    
    #Create sheet
    sheet = workbook.create_sheet(title='variable_importance')
    
    #Set border style
    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

    # Write the column headers in row 2
    start_row = 2
    start_col = 2
    for c_idx, col in enumerate(df.columns):
        cell = sheet.cell(row=start_row, column=start_col + c_idx, value=col)
        cell.border = thin_border

    # Write the DataFrame data and row names
    for r_idx, (idx, row) in enumerate(df.iterrows()):
        # Write row name in column A (assuming start_col=1 for column A)
        cell = sheet.cell(row=start_row + 1 + r_idx, column=1, value=idx)
        cell.border = thin_border
        # Write data values starting from column B
        for c_idx, value in enumerate(row):
            cell = sheet.cell(row=start_row + 1 + r_idx, column=start_col + c_idx, value=value)
            cell.border = thin_border

    # Define the conditional formatting rule for a 3-color scale
    color_scale_rule = ColorScaleRule(
        start_type='min', start_color='FF0000',
        mid_type='percentile', mid_value=50, mid_color='FFFFFF',
        end_type='max', end_color='00FF00'
    )

    # Apply the conditional formatting to each column in the DataFrame
    for c_idx, col in enumerate(df.columns):
        col_letter = chr(ord('B') + c_idx)  # Convert column index to letter (e.g., 'B' for start_col=2)
        col_range = f"{col_letter}{start_row + 1}:{col_letter}{start_row + 1 + len(df)}"
        sheet.conditional_formatting.add(col_range, color_scale_rule)

    #center everything
    alignment = Alignment(horizontal='center', vertical='center')
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = alignment

    #no gridlines
    sheet.sheet_view.showGridLines = False
    
    #autofit column width
    for col in range(1, 9):  # Columns B to F
        column_letter = get_column_letter(col)
        max_length = 0
        for row in range(1, sheet.max_row + 1):
                cell_value = sheet[f'{column_letter}{row}'].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
        adjusted_width = (max_length) * 1.2  # Adjusting factor for better fit
        sheet.column_dimensions[column_letter].width = adjusted_width

    #autofit row height
    for row in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 15  # Adjust the height as needed   

    #add in plot
    graph_path = variable_importance_graph(df,result_path)
    img = Image(graph_path)
    
    last_col = sheet.max_column
    insert_col = last_col + 2
    letter = get_column_letter(insert_col)
    
    sheet.add_image(img,f'{letter}2')
    
def heatmap(workbook,df,test_set_names, X_cols,result_path):
    """
    Creates the variable importance heatmap.
    """    
    #Create sheet
    sheet = workbook.create_sheet(title='heatmap')
    
    #add in plot
    graph_path = heatmap_graph(df,test_set_names,X_cols,result_path)
    img = Image(graph_path)

    sheet.add_image(img,'A1')

def generate_workbook(model_analysis,result_path,PredictionResults,test_set_names,X_cols,y_actuals):
    """
    Creates the excel workbook
    """
    workbook = openpyxl.Workbook()
    
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    

    
    create_y_actual_means(workbook, model_analysis[0],result_path)
    create_co_occurrence(workbook, model_analysis[1],result_path)
    betas_and_t_statistics(workbook, model_analysis[2])
    variable_importance(workbook,model_analysis[3],result_path)
    test_set_statistics(workbook,PredictionResults,test_set_names,y_actuals)
    heatmap(workbook,PredictionResults.combi_compound,test_set_names,X_cols,result_path)
    
    
    workbook.save(f"{result_path}\\results.xlsx")
    
    os.startfile(f"{result_path}\\results.xlsx")
    
    